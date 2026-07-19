"""
export_to_xlsx.py
-----------------
Exports master.json to a clean Excel spreadsheet.

Usage:
    python pipeline/export_to_xlsx.py
    python pipeline/export_to_xlsx.py --input data/master.json --output data/master.xlsx
    python pipeline/export_to_xlsx.py --status complete  # only complete records
"""

import json
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DEFAULT_INPUT  = Path("data/master.json")
from datetime import datetime
DEFAULT_OUTPUT = Path(f"data/master_{datetime.now().strftime('%Y%m%d')}.xlsx")

# Columns to export and their display names
COLUMNS = [
    ("firm_name",              "Firm Name"),
    ("firm_type",              "Type"),
    ("hq_state",               "State"),
    ("hq_city",                "City"),
    ("website",                "Website"),
    ("us_investments",         "US Investments"),
    ("revenue_min",            "Revenue Min ($M)"),
    ("revenue_max",            "Revenue Max ($M)"),
    ("ebitda_min",             "EBITDA Min ($M)"),
    ("ebitda_max",             "EBITDA Max ($M)"),
    ("enterprise_value_min",   "EV Min ($M)"),
    ("enterprise_value_max",   "EV Max ($M)"),
    ("check_size_min",         "Check Min ($M)"),
    ("check_size_max",         "Check Max ($M)"),
    ("deal_types",             "Deal Types"),
    ("sector_tier1",           "Sector Tier 1"),
    ("sector_tier2",           "Sector Tier 2"),
    ("sector_tier3",           "Sector Tier 3"),
    ("sector_exclusions",      "Exclusions"),
    ("aum_usd_millions",       "AUM ($M)"),
    ("activity_tier",          "Activity"),
    ("last_known_deal_date",   "Last Deal"),
    ("fund_name",              "Fund Name"),
    ("confidence",             "Confidence"),
    ("needs_review",           "Needs Review"),
    ("last_checked_date",      "Last Checked"),
    ("status",                 "Status"),
    ("notes",                  "Notes"),
]

# Color coding
COLORS = {
    "header_bg":    "1F3864",  # Dark navy
    "header_fg":    "FFFFFF",  # White
    "high_conf":    "C6EFCE",  # Green
    "medium_conf":  "FFEB9C",  # Yellow
    "low_conf":     "FFC7CE",  # Red
    "needs_review": "FCE4D6",  # Orange
    "alt_row":      "F2F2F2",  # Light grey
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def format_value(key: str, value) -> str:
    """Format a value for display in Excel."""
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def get_row_color(record: dict) -> str | None:
    """Return background color based on confidence/review status."""
    if record.get("needs_review"):
        return COLORS["needs_review"]
    conf = record.get("confidence", "")
    if conf == "High":
        return COLORS["high_conf"]
    if conf == "Medium":
        return COLORS["medium_conf"]
    if conf == "Low":
        return COLORS["low_conf"]
    return None


# ---------------------------------------------------------------------------
# Main export
# ---------------------------------------------------------------------------

def export(input_path: Path, output_path: Path, status_filter: str | None = None):
    # Load master
    with open(input_path, encoding="utf-8") as f:
        records = json.load(f)

    # Filter by status if requested
    if status_filter:
        records = [r for r in records if r.get("status") == status_filter]

    # Sort: complete first, then by firm name
    records.sort(key=lambda r: (r.get("status") != "complete", r.get("firm_name", "")))

    print(f"Exporting {len(records)} records to {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Firms"

    # --- Header row ---
    header_font    = Font(name="Arial", bold=True, color=COLORS["header_fg"], size=10)
    header_fill    = PatternFill("solid", start_color=COLORS["header_bg"])
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border    = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, (field, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border

    ws.row_dimensions[1].height = 30

    # --- Data rows ---
    data_font  = Font(name="Arial", size=9)
    data_align = Alignment(vertical="top", wrap_text=False)
    notes_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, record in enumerate(records, start=2):
        row_color = get_row_color(record)
        row_fill  = PatternFill("solid", start_color=row_color) if row_color else (
                    PatternFill("solid", start_color=COLORS["alt_row"]) if row_idx % 2 == 0 else None
                    )

        for col_idx, (field, label) in enumerate(COLUMNS, start=1):
            value     = record.get(field)
            formatted = format_value(field, value)
            cell      = ws.cell(row=row_idx, column=col_idx, value=formatted)
            cell.font = data_font
            cell.alignment = notes_align if field == "notes" else data_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 15

    # --- Column widths ---
    col_widths = {
        "firm_name":            30,
        "firm_type":            10,
        "hq_state":              6,
        "hq_city":              12,
        "website":              28,
        "us_investments":       10,
        "revenue_min":          12,
        "revenue_max":          12,
        "ebitda_min":           12,
        "ebitda_max":           12,
        "enterprise_value_min": 10,
        "enterprise_value_max": 10,
        "check_size_min":       10,
        "check_size_max":       10,
        "deal_types":           22,
        "sector_tier1":         28,
        "sector_tier2":         28,
        "sector_tier3":         28,
        "sector_exclusions":    20,
        "aum_usd_millions":     10,
        "activity_tier":        10,
        "last_known_deal_date": 12,
        "fund_name":            20,
        "confidence":           10,
        "needs_review":          9,
        "last_checked_date":    12,
        "status":               12,
        "notes":                45,
    }

    for col_idx, (field, _) in enumerate(COLUMNS, start=1):
        width = col_widths.get(field, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto filter ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "ICS Pipeline Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)

    summary_data = [
        ("Export Date",     datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Records",   len(records)),
        ("Complete",        sum(1 for r in records if r.get("status") == "complete")),
        ("Pending",         sum(1 for r in records if r.get("status") == "pending")),
        ("Needs Review",    sum(1 for r in records if r.get("needs_review"))),
        ("High Confidence", sum(1 for r in records if r.get("confidence") == "High")),
        ("Med Confidence",  sum(1 for r in records if r.get("confidence") == "Medium")),
        ("Low Confidence",  sum(1 for r in records if r.get("confidence") == "Low")),
        ("Active Firms",    sum(1 for r in records if r.get("activity_tier") == "Active")),
        ("Slowing Firms",   sum(1 for r in records if r.get("activity_tier") == "Slowing")),
        ("Dormant Firms",   sum(1 for r in records if r.get("activity_tier") == "Dormant")),
    ]

    for i, (label, value) in enumerate(summary_data, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=i, column=2, value=value).font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    # --- Save ---
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    print(f"✓ Saved: {output_path}")
    print(f"  {len(records)} firms | {sum(1 for r in records if r.get('status') == 'complete')} complete")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export master.json to Excel")
    parser.add_argument("--input",  default=str(DEFAULT_INPUT),  help="Path to master.json")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Path for output .xlsx")
    parser.add_argument("--status", default=None, help="Filter by status (e.g. complete)")
    args = parser.parse_args()

    export(
        input_path    = Path(args.input),
        output_path   = Path(args.output),
        status_filter = args.status,
    )